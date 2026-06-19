import sys
import json
import io
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def auto_size_columns(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

def main():
    try:
        input_data = sys.stdin.read().strip()
        if not input_data:
            # Send empty workbook
            wb = Workbook()
            output = io.BytesIO()
            wb.save(output)
            sys.stdout.buffer.write(output.getvalue())
            return

        records = json.loads(input_data)
        
        if not records:
            wb = Workbook()
            wb.save(sys.stdout.buffer)
            return

        # Prepare Student Info
        student_name = records[0]['Student Name']
        reg_no = records[0]['Register No']

        # Group records by Month-Year
        # Date format expected: "YYYY-MM-DD"
        months = {}
        for r in records:
            date_str = r['Date']
            try:
                dt = datetime.datetime.strptime(date_str, "%Y-%m-%d")
                month_key = dt.strftime("%b %Y") # e.g. "Jun 2025"
                sort_key = dt.strftime("%Y-%m")
            except:
                month_key = "Unknown"
                sort_key = "0000-00"
                
            if sort_key not in months:
                months[sort_key] = {'label': month_key, 'records': []}
            months[sort_key]['records'].append(r)

        wb = Workbook()
        
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell_font = Font(name='Arial', size=11)
        bold_font = Font(name='Arial', size=11, bold=True)
        
        # --- Summary Sheet ---
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        ws_summary.append(["Student Name:", student_name])
        ws_summary.append(["Register No:", reg_no])
        ws_summary.append([]) # Blank row
        
        ws_summary.cell(row=1, column=1).font = bold_font
        ws_summary.cell(row=2, column=1).font = bold_font
        
        sum_headers = ['Month', 'Total Present', 'Total Absent', 'Attendance %']
        ws_summary.append(sum_headers)
        
        header_row = 4
        for col_num, header in enumerate(sum_headers, 1):
            cell = ws_summary.cell(row=header_row, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            
        # Sort months chronologically
        sorted_month_keys = sorted(list(months.keys()))
        
        current_row = header_row + 1
        for m_key in sorted_month_keys:
            m_label = months[m_key]['label']
            ws_summary.cell(row=current_row, column=1, value=m_label).font = cell_font
            
            # Count Present and Absent
            present_c = sum(1 for x in months[m_key]['records'] if x['Status'] == 'PRESENT')
            absent_c = sum(1 for x in months[m_key]['records'] if x['Status'] == 'ABSENT')
            
            ws_summary.cell(row=current_row, column=2, value=present_c).font = cell_font
            ws_summary.cell(row=current_row, column=3, value=absent_c).font = cell_font
            
            # Percentage
            perc_cell = ws_summary.cell(row=current_row, column=4, value=f"=IF(B{current_row}+C{current_row}>0, B{current_row}/(B{current_row}+C{current_row}), 0)")
            perc_cell.font = cell_font
            perc_cell.number_format = '0.00%'
            current_row += 1

        # Overall Totals
        ws_summary.cell(row=current_row, column=1, value="OVERALL TOTALS").font = bold_font
        ws_summary.cell(row=current_row, column=2, value=f"=SUM(B5:B{current_row-1})").font = bold_font
        ws_summary.cell(row=current_row, column=3, value=f"=SUM(C5:C{current_row-1})").font = bold_font
        total_perc = ws_summary.cell(row=current_row, column=4, value=f"=IF(B{current_row}+C{current_row}>0, B{current_row}/(B{current_row}+C{current_row}), 0)")
        total_perc.font = bold_font
        total_perc.number_format = '0.00%'

        auto_size_columns(ws_summary)
        
        # --- Month Sheets ---
        for m_key in sorted_month_keys:
            m_label = months[m_key]['label']
            
            # Sheet names in Excel cannot be longer than 31 chars and no special chars.
            safe_title = m_label[:31] 
            ws_month = wb.create_sheet(title=safe_title)
            
            date_headers = ['Date', 'Section', 'Status']
            ws_month.append(date_headers)
            
            for col_num, header in enumerate(date_headers, 1):
                cell = ws_month.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill

            m_records = months[m_key]['records']
            m_records.sort(key=lambda x: x['Date'])
            
            for i, r in enumerate(m_records, start=2):
                ws_month.cell(row=i, column=1, value=r['Date']).font = cell_font
                ws_month.cell(row=i, column=2, value=r['Section']).font = cell_font
                ws_month.cell(row=i, column=3, value=r['Status']).font = cell_font
                
            auto_size_columns(ws_month)

        # Output workbook to stdout
        output = io.BytesIO()
        wb.save(output)
        sys.stdout.buffer.write(output.getvalue())

    except Exception as e:
        sys.stderr.write(str(e))
        sys.exit(1)

if __name__ == "__main__":
    main()
