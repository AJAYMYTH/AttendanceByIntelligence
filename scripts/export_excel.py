import sys
import json
import io
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def auto_size_columns(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

def main():
    try:
        input_data = sys.stdin.read().strip()
        if not input_data:
            # Send empty workbook
            wb = Workbook()
            output = io.BytesIO()
            wb.save(output)
            sys.stdout.buffer.write(output.getvalue())
            return

        records = json.loads(input_data)
        
        if not records:
            wb = Workbook()
            wb.save(sys.stdout.buffer)
            return

        # Group records by Month-Year
        months_dict = {}
        for r in records:
            date_str = r['Date']
            try:
                dt = datetime.datetime.strptime(date_str, "%Y-%m-%d")
                month_label = dt.strftime("%b %Y") # e.g. "Jun 2025"
                sort_key = dt.strftime("%Y-%m")
            except:
                month_label = "Unknown Data"
                sort_key = "0000-00"
            
            if sort_key not in months_dict:
                months_dict[sort_key] = {'label': month_label, 'records': []}
            months_dict[sort_key]['records'].append(r)

        # Extract unique students
        students_dict = {}
        for r in records:
            sig = (r['Student Name'], r['Register No'], r['Section'])
            if sig not in students_dict:
                students_dict[sig] = { 
                    'name': r['Student Name'], 
                    'reg': r['Register No'], 
                    'sec': r['Section'],
                    'present_count': 0,
                    'absent_count': 0
                }
            if r['Status'] == 'PRESENT':
                students_dict[sig]['present_count'] += 1
            elif r['Status'] == 'ABSENT':
                students_dict[sig]['absent_count'] += 1

        students = sorted(list(students_dict.values()), key=lambda x: x['name'])

        wb = Workbook()
        
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell_font = Font(name='Arial', size=11)
        bold_font = Font(name='Arial', size=11, bold=True)
        
        # --- Summary Sheet ---
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        sum_headers = ['Student Name', 'Register No', 'Section', 'Total Present', 'Total Absent', 'Attendance %']
        ws_summary.append(sum_headers)
        
        for col_num, header in enumerate(sum_headers, 1):
            cell = ws_summary.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            
        for i, s in enumerate(students, start=2):
            ws_summary.cell(row=i, column=1, value=s['name']).font = cell_font
            ws_summary.cell(row=i, column=2, value=s['reg']).font = cell_font
            ws_summary.cell(row=i, column=3, value=s['sec']).font = cell_font
            
            # Using raw counts instead of buggy cross-sheet formulas, to ensure it always correctly displays
            ws_summary.cell(row=i, column=4, value=s['present_count']).font = cell_font
            ws_summary.cell(row=i, column=5, value=s['absent_count']).font = cell_font
            
            # Percentage Formula
            perc_cell = ws_summary.cell(row=i, column=6, value=f"=IF(D{i}+E{i}>0, D{i}/(D{i}+E{i}), 0)")
            perc_cell.font = cell_font
            perc_cell.number_format = '0.00%'

        last_student_row = len(students) + 1
        summary_row = last_student_row + 1
        
        ws_summary.cell(row=summary_row, column=1, value="TOTALS/AVERAGES").font = bold_font
        ws_summary.cell(row=summary_row, column=4, value=f"=SUM(D2:D{last_student_row})").font = bold_font
        ws_summary.cell(row=summary_row, column=5, value=f"=SUM(E2:E{last_student_row})").font = bold_font
        perc_total = ws_summary.cell(row=summary_row, column=6, value=f"=AVERAGE(F2:F{last_student_row})")
        perc_total.font = bold_font
        perc_total.number_format = '0.00%'

        auto_size_columns(ws_summary)
        
        # --- Month Sheets ---
        sorted_month_keys = sorted(list(months_dict.keys()))
        for m_key in sorted_month_keys:
            m_label = months_dict[m_key]['label']
            
            safe_title = m_label[:31]
            ws_month = wb.create_sheet(title=safe_title)
            
            date_headers = ['Date', 'Student Name', 'Register No', 'Section', 'Status']
            ws_month.append(date_headers)
            
            for col_num, header in enumerate(date_headers, 1):
                cell = ws_month.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill

            date_records = months_dict[m_key]['records']
            # Sort by Date then Name
            date_records.sort(key=lambda x: (x['Date'], x['Student Name']))
            
            for i, r in enumerate(date_records, start=2):
                ws_month.cell(row=i, column=1, value=r['Date']).font = cell_font
                ws_month.cell(row=i, column=2, value=r['Student Name']).font = cell_font
                ws_month.cell(row=i, column=3, value=r['Register No']).font = cell_font
                ws_month.cell(row=i, column=4, value=r['Section']).font = cell_font
                ws_month.cell(row=i, column=5, value=r['Status']).font = cell_font
                
            last_row = len(date_records) + 1
            tot_row = last_row + 1
            
            ws_month.cell(row=tot_row, column=1, value="TOTALS").font = bold_font
            ws_month.cell(row=tot_row, column=5, value=f"=\"P: \"&COUNTIF(E2:E{last_row}, \"PRESENT\")&\", A: \"&COUNTIF(E2:E{last_row}, \"ABSENT\")").font = bold_font
            
            auto_size_columns(ws_month)

        # Output workbook to stdout
        output = io.BytesIO()
        wb.save(output)
        sys.stdout.buffer.write(output.getvalue())

    except Exception as e:
        sys.stderr.write(str(e))
        sys.exit(1)

if __name__ == "__main__":
    main()
